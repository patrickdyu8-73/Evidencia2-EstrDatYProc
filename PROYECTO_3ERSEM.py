from datetime import datetime, timedelta
import sqlite3
from sqlite3 import Error
import sys

hoy = datetime.now().date() 


try:
    with sqlite3.connect('espacios_coworking.db', detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
        cursor = conn.cursor()

        cursor.execute('CREATE TABLE IF NOT EXISTS clientes \
                        (id_cliente INTEGER PRIMARY KEY, \
                        nombre TEXT NOT NULL, \
                        apellido TEXT NOT NULL);')
        

        cursor.execute('CREATE TABLE IF NOT EXISTS salas \
                        (id_sala INTEGER PRIMARY KEY, \
                        nombre_sala TEXT NOT NULL, \
                        cupo_sala INTEGER NOT NULL);')
        

        cursor.execute('CREATE TABLE IF NOT EXISTS reservaciones \
                        (folio_reservacion INTEGER PRIMARY KEY, \
                        nombre_reservacion TEXT NOT NULL, \
                        turno_reservacion TEXT NOT NULL, \
                        fecha_reservacion TIMESTAMP NOT NULL, \
                        estatus_reservacion TEXT DEFAULT "Activa", \
                        id_cliente INTEGER, \
                        id_sala INTEGER, \
                        FOREIGN KEY (id_cliente) REFERENCES clientes(id_cliente), \
                        FOREIGN KEY (id_sala) REFERENCES salas(id_sala));')
        
except Error as e:
    print(e)
except:
    print(f'Se produjo un error: {sys.exc_info()[0]}')





while True:
    print('\n************ MENÚ ************')
    print('1. REGISTRAR RESERVACION DE UNA SALA')
    print('2. EDITAR EL NOMBRE DE UN EVENTO DE UNA RESERVACION YA HECHA')
    print('3. CONSULTAR RESERVACIONES EXISTENTES')
    print('4. CANCELAR UNA RESERVA')
    print('5. REGISTRAR UN CLIENTE')
    print('6. REGISTRAR UNA SALA')
    print('7. SALIR') 

    opcion = input('Ingrese una opción: ')


    if opcion not in '1234567':
        print('\nRespuesta no valida')
        continue
    

    if opcion == '1':
        try:
            with sqlite3.connect('espacios_coworking.db') as conn:
                cursor = conn.cursor()
                cursor.execute('SELECT id_cliente FROM clientes')
                cliente_registrados = cursor.fetchall()
        except Error as e:
            print(e)
        except:
            print(f'Se produjo un error: {sys.exc_info()[0]}')
        
        if cliente_registrados:
            while True:
                try:
                    with sqlite3.connect('espacios_coworking.db') as conn:
                        cursor = conn.cursor()
                        cursor.execute('SELECT id_cliente, apellido, nombre FROM clientes\
                                        ORDER BY apellido')
                        clientes_ordenados = cursor.fetchall()

                        if clientes_ordenados:
                            print('\n********** CLIENTES REGISTRADOS **********')
                            print(f'{'ID':<10} {'APELLIDO':<20} {'NOMBRE'}')
                            print('='*42)
                            for id_cliente,apellido_cliente,nombre_cliente in clientes_ordenados:
                                print(f'{id_cliente:<10} {apellido_cliente:<20} {nombre_cliente}')
                        else:
                            print('No hay clientes registrados')
                except Error as e:
                    print(e)
                except:
                    print(f'Se produjo un error: {sys.exc_info()[0]}')


                id_cliente_buscar = input('\nDigite su ID ("X" para salir): ').strip().upper()
                if id_cliente_buscar == 'X':
                    print('Saliendo...')
                    break
                if not id_cliente_buscar.isdigit():
                    print('Id invalido')
                    continue
                else:
                    id_cliente_buscar = int(id_cliente_buscar)
                    try:
                        with sqlite3.connect('espacios_coworking.db') as conn:
                            cursor = conn.cursor()
                            buscar_cliente = {'clave':id_cliente_buscar}
                            cursor.execute('SELECT id_cliente FROM clientes \
                                           WHERE id_cliente = :clave',buscar_cliente)
                            cliente_buscado = cursor.fetchone()
                            if cliente_buscado:
                                print('ID encontrado')
                            else:
                                print('ID no encontrado')
                                continue
                    except Error as e:
                        print(e)
                    except:
                        print(f'Se produjo un error: {sys.exc_info()[0]}')
                
                fecha_minima = hoy + timedelta(days=2)
                while True:
                    fecha_usuario_str=input("\nINGRESE LA FECHA DE RESERVA (MM-DD-AAAA): ").strip()
                    try:
                        fecha_usuario = datetime.strptime(fecha_usuario_str, "%m-%d-%Y").date() 

                        if fecha_usuario >= fecha_minima:
                            if fecha_usuario.weekday() == 6:
                                lunes_siguiente = fecha_usuario + timedelta(days=1)
                                print(f'Error. No se permiten reservaciones en domingo ({fecha_usuario.strftime("%m-%d-%Y")}).')
                                print(f'Sugerencia: Puede reservar el lunes {lunes_siguiente.strftime("%m-%d-%Y")}.')
                                continue
                            else:
                                print("La fecha es válida para reserva")
                                break
                        else:
                            print(f'Error. La fecha debe de ser por lo menos 2 días posteriores a hoy {hoy}')
                            continue

                    except ValueError:
                        print("Formato de fecha invalido")
                        continue
                

                while True:
                    print('TURNOS: "Matutino"(M) | "Vespertino"(V) | "Nocturno"(N)')
                    seleccionar_turno = input("Ingrese turno el turno que desea (M/V/N): ").upper().strip()
                    if seleccionar_turno:
                        if seleccionar_turno == 'M':
                            turno = 'Matutino'
                        elif seleccionar_turno == 'V':
                            turno = 'Vespertino'
                        elif seleccionar_turno == 'N':
                            turno = 'Nocturno'
                        else:
                            print('Turno invalido')
                            continue
                        print('Turno registrado')
                        break
                    else:
                        print('Error. Seleccione un turno valido')
                        continue


                try:
                    with sqlite3.connect('espacios_coworking.db') as conn:
                        cursor = conn.cursor()
                        cursor.execute('SELECT id_sala, nombre_sala, cupo_sala FROM salas')
                        salas_totales = cursor.fetchall()

                        if not salas_totales:
                            print('No hay salas registradas aún.')
                            continue


                        datos_buscar = {'fecha': fecha_usuario, 'turno': turno}
                        cursor.execute('SELECT id_sala FROM reservaciones \
                                        WHERE fecha_reservacion = :fecha \
                                        AND turno_reservacion = :turno', datos_buscar)
                        salas_ocupadas = [fila[0] for fila in cursor.fetchall()]


                        salas_disponibles = [(id_sala, nombre_sala, cupo_sala)
                                                for id_sala, nombre_sala, cupo_sala in salas_totales
                                                if id_sala not in salas_ocupadas]

                        if salas_disponibles:
                            print('\n********** SALAS DISPONIBLES **********')
                            print(f'{"ID":<5}{"NOMBRE":<20}{"CUPO":<10}')
                            print('=' * 35)
                            for id_sala, nombre_sala, cupo_sala in salas_disponibles:
                                print(f'{id_sala:<5}{nombre_sala:<20}{cupo_sala:<10}')
                        else:
                            print('No hay salas disponibles para esa fecha y turno.')
                            continue
                except Error as e:
                    print(e)
                except:
                    print(f'Se produjo un error: {sys.exc_info()[0]}')


                salas_ids = [id_sala for id_sala, nombre_sala, cupo_sala in salas_disponibles]

                while True:
                    try:
                        sala_seleccionada = int(input("\nIngrese el ID de la sala que desea reservar: "))
                    except ValueError:
                        print("Debe ingresar un número de sala válido.")
                        continue

                    if sala_seleccionada not in salas_ids:
                        print("La sala seleccionada no está disponible.")
                        continue
                    else:
                        print("Sala seleccionada correctamente.")
                        break


                while True:
                    nombre_reserva = input('\nIngrese el nombre del evento o reservación: ').strip().title()
                    if not nombre_reserva:
                        print('Error. El nombre no puede estar vacío.')
                        continue
                    else:
                        break


                reservacion = (nombre_reserva, turno, fecha_usuario, id_cliente_buscar, sala_seleccionada)

                try:
                    with sqlite3.connect('espacios_coworking.db') as conn:
                        cursor = conn.cursor()
                        cursor.execute('INSERT INTO reservaciones (nombre_reservacion, turno_reservacion, \
                                    fecha_reservacion, id_cliente, id_sala) VALUES (?,?,?,?,?)', reservacion) 
                        folio = cursor.lastrowid
                        print(f'\nReservación registrada exitosamente con el folio #{folio}.')
                        print(f'Fecha: {fecha_usuario.strftime("%m-%d-%Y")}')
                        break
                except Error as e:
                    print(e)
                except:
                    print(f'Se produjo un error: {sys.exc_info()[0]}')
        
        else:
            print('No hay clientes registrados')
            continue




    if opcion == '2':
        while True:
            try:
                fecha_inicio_str = input("\nIngrese fecha de inicio (MM-DD-AAAA) ('X' para salir): ").strip().upper()
                if fecha_inicio_str == 'X':
                    print('Saliendo...')
                    break
                fecha_fin_str = input("Ingrese fecha final (MM-DD-AAAA): ").strip()
                
                fecha_inicio = datetime.strptime(fecha_inicio_str, "%m-%d-%Y").date()
                fecha_fin = datetime.strptime(fecha_fin_str, "%m-%d-%Y").date()
                if fecha_fin < fecha_inicio:
                    print("Error. La fecha final no puede ser anterior a la fecha de inicio.")
                    continue
            except ValueError:
                print("Formato de fecha inválido")
                continue

            fechas = (fecha_inicio, fecha_fin)
            
            try:
                with sqlite3.connect('espacios_coworking.db') as conn:
                    cursor = conn.cursor()
                    cursor.execute('SELECT folio_reservacion, nombre_reservacion, \
                                    turno_reservacion, fecha_reservacion \
                                    FROM reservaciones \
                                    WHERE fecha_reservacion BETWEEN ? AND ? \
                                    AND estatus_reservacion == "Activa"', fechas)
                    cancelacion_reservas = cursor.fetchall()

                if cancelacion_reservas:
                    print('\n***** RESERVACIONES EN EL RANGO DE FECHAS *****')
                    print(f'{"FOLIO":<10}{"NOMBRE":<20}{"TURNO":<15}{"FECHA"}')
                    print('='*55)
                    
                    for folio, nombre, turno, fecha in cancelacion_reservas:
                        try:
                            fecha_formateada = datetime.fromisoformat(fecha).strftime('%m-%d-%Y')
                        except ValueError:
                            fecha_formateada = fecha

                        print(f'{folio:<10}{nombre:<20}{turno:<15}{fecha_formateada}')
                
                else: 
                    print('No hay reservaciones en este rango de fechas')
                    continue
            
            except Error as e:
                print(f"Error de base de datos: {e}")
                continue
            except Exception as e:
                print(f'Se produjo un error {e}')
                continue

            folio_seleccionado = input("\nIngrese el folio de la reserva que desea modificar: ").strip().upper()
            folios_validos = [fila[0] for fila in cancelacion_reservas]
            try:
                folio_seleccionado = int(folio_seleccionado)
            except ValueError:
                print('Ingrese un número de folio valido')
                continue
            if folio_seleccionado not in folios_validos:
                print('Folio no valido')
                continue

            nuevo_evento = input("Ingrese el nuevo nombre del evento: ").strip()
            if not nuevo_evento:
                print("El nombre del evento no puede estar vacío.")
                continue       
            
            editar_datos = (nuevo_evento, folio_seleccionado)
            try:
                with sqlite3.connect('espacios_coworking.db') as conn:
                    cursor = conn.cursor()
                    cursor.execute('UPDATE reservaciones \
                                    SET nombre_reservacion = ? \
                                    WHERE folio_reservacion = ?', editar_datos)
                    conn.commit()
                    print('Nombre de reservacion actualizado')
                    break

            except Error as e:
                print(e)
            except:
                print(f'Se produjo un error: {sys.exc_info()[0]}')
            


    if opcion == '3':
        while True:
            fecha_consulta_str = input("\nIngrese la fecha a consultar (MM-DD-AAAA) ('X' para salir): ").strip()
            
            if fecha_consulta_str.upper() == 'X':
                print('Saliendo...')
                break

            if fecha_consulta_str:
                try:
                    fecha_consulta = datetime.strptime(fecha_consulta_str, "%m-%d-%Y").date()
                except ValueError:
                    print("Formato de fecha invalido")
                    continue

            else:
                hoy = datetime.now().date()
                fecha_consulta = hoy
            
            datos_reservacion = []
            try:
                with sqlite3.connect('espacios_coworking.db') as conn:
                    cursor = conn.cursor()
                    cursor.execute('SELECT c.nombre, c.apellido, s.id_sala, r.nombre_reservacion, r.turno_reservacion \
                                FROM reservaciones AS r\
                                INNER JOIN clientes AS c ON r.id_cliente = c.id_cliente\
                                INNER JOIN salas AS s ON r.id_sala = s.id_sala\
                                WHERE r.fecha_reservacion = ? \
                                AND estatus_reservacion = "Activa"', (fecha_consulta,))
                    datos_reservacion = cursor.fetchall()

                    if datos_reservacion:
                        print('*'*75)
                        print(f'** {"REPORTE DE RESERVACIONES PARA EL DÍA":^59}{fecha_consulta} **')
                        print('*'*75)
                        print(f'{'SALA':<10} {'CLIENTE':<30} {'EVENTO':<20} {'TURNO'}')
                        print('='*75)
                        for nombre, apellido, id_sala, evento, turno in datos_reservacion:
                            cliente = f'{apellido} {nombre}'
                            print(f'{id_sala:<10} {cliente:<30} {evento:<20} {turno}')
                        print(f'** {"FIN DEL REPORTE":^69} **')
                    else:
                        print('No hay reservaciones para esta fecha')
                        continue
                
            except Error as e:
                print(e)
            except:
                print(f'Se produjo un error: {sys.exc_info()[0]}')

            if datos_reservacion:
                while True:
                    opcion_exportar = input('Desea exportar el reporte tabular? (S/N): ').strip().upper()
                    if opcion_exportar == 'S':
                        print('a) CSV')
                        print('b) JSON')
                        print('c) EXCEL')
                        formato_exportar = input('Seleccione el formato de exportacion: ').strip().lower()
                        
                        if formato_exportar == 'a':
                            import csv
                            try:
                                with sqlite3.connect('espacios_coworking.db') as conn:
                                    cursor = conn.cursor()
                                    cursor.execute('SELECT s.id_sala, c.apellido || " " || c.nombre AS cliente, \
                                                    r.nombre_reservacion, r.turno_reservacion \
                                                    FROM reservaciones AS r\
                                                    INNER JOIN clientes AS c ON r.id_cliente = c.id_cliente\
                                                    INNER JOIN salas AS s ON r.id_sala = s.id_sala\
                                                    WHERE r.fecha_reservacion = ?\
                                                    AND estatus_reservacion = "Activa"', (fecha_consulta,))
                                    datos = cursor.fetchall()
                                    columnas = [desc[0] for desc in cursor.description]

                                    if datos:
                                        with open('reservaciones.csv', 'w', newline='', encoding='utf-8') as archivo_csv:
                                            escritor = csv.writer(archivo_csv)
                                            escritor.writerow(columnas)
                                            escritor.writerows(datos)
                                        print('Archivo "reservaciones.csv" exportado correctamente.')
                                    else:
                                        print('No hay datos para exportar.')
                            except Error as e:
                                print(e)
                            except Exception as e:
                                print(f'Se produjo un error al exportar a CSV: {e}')

                        elif formato_exportar == 'b':
                            import json
                            try:
                                with sqlite3.connect('espacios_coworking.db') as conn:
                                    cursor = conn.cursor()
                                    cursor.execute('SELECT s.id_sala, c.apellido || " " || c.nombre AS cliente, \
                                                    r.nombre_reservacion, r.turno_reservacion \
                                                    FROM reservaciones AS r\
                                                    INNER JOIN clientes AS c ON r.id_cliente = c.id_cliente\
                                                    INNER JOIN salas AS s ON r.id_sala = s.id_sala\
                                                    WHERE r.fecha_reservacion = ?\
                                                    AND estatus_reservacion = "Activa"', (fecha_consulta,))
                                    datos = cursor.fetchall()
                                    columnas = [desc[0] for desc in cursor.description]

                                    if datos:
                                        lista_dicts = [dict(zip(columnas, fila)) for fila in datos]
                                        with open('reservaciones.json', 'w', encoding='utf-8') as archivo_json:
                                            json.dump(lista_dicts, archivo_json, ensure_ascii=False, indent=4)
                                        print('Archivo "reservaciones.json" exportado correctamente.')
                                    else:
                                        print('No hay datos para exportar.')
                            except Error as e:
                                    print(e)
                            except Exception as e:
                                    print(f'Se produjo un error al exportar a JSON: {e}')

                        elif formato_exportar == 'c':
                            from openpyxl import Workbook
                            from openpyxl.styles import Font, Alignment, Border, Side
                            from openpyxl.utils import get_column_letter

                            try:
                                with sqlite3.connect('espacios_coworking.db') as conn:
                                    cursor = conn.cursor()
                                    
                                    cursor.execute('SELECT r.folio_reservacion, c.nombre || " " || c.apellido AS cliente,\
                                                    s.nombre_sala, r.nombre_reservacion, r.turno_reservacion, r.fecha_reservacion\
                                                    FROM reservaciones AS r\
                                                    INNER JOIN clientes AS c ON r.id_cliente = c.id_cliente\
                                                    INNER JOIN salas AS s ON r.id_sala = s.id_sala\
                                                    WHERE r.fecha_reservacion = ?\
                                                    AND estatus_reservacion = "Activa"\
                                                    ORDER BY r.fecha_reservacion', (fecha_consulta,))
                                    datos = cursor.fetchall()
                                    columnas = [desc[0] for desc in cursor.description]

                                    if datos:
                                        libro = Workbook()
                                        hoja = libro.active
                                        hoja.title = "Reservaciones"
                                        
                                        
                                        hoja.merge_cells('A1:F1')
                                        hoja['A1'] = "REPORTE DE RESERVACIONES"
                                        hoja['A1'].font = Font(bold=True, size=14)
                                        hoja['A1'].alignment = Alignment(horizontal='center', vertical='center')

                                        encabezado_font = Font(bold=True)
                                        borde_grueso = Border(bottom=Side(style='thick'))
                                        alineacion_centrada = Alignment(horizontal='center', vertical='center')

                                        for num_col, encabezado in enumerate(columnas, 1):
                                            celda = hoja.cell(row=3, column=num_col, value=encabezado.upper())
                                            celda.font = encabezado_font
                                            celda.border = borde_grueso
                                            celda.alignment = alineacion_centrada

                                        for fila_num, fila_datos in enumerate(datos, 4):
                                            for num_col, valor in enumerate(fila_datos, 1):
                                                celda = hoja.cell(row=fila_num, column=num_col, value=valor)
                                                celda.alignment = alineacion_centrada

                                        for x in range(1, len(columnas) + 1):
                                            letra_col = get_column_letter(x)
                                            longitud_max = 0
                                            for fila in range(3, hoja.max_row + 1):
                                                valor_celda = hoja.cell(row=fila, column=x).value
                                                try:
                                                    if valor_celda and len(str(valor_celda)) > longitud_max:
                                                        longitud_max = len(str(valor_celda))
                                                except:
                                                    pass
                                            hoja.column_dimensions[letra_col].width = longitud_max + 2

                                        libro.save(f"reservaciones_{fecha_consulta}.xlsx")
                                        print(f"Reporte 'reservaciones_{fecha_consulta}.xlsx' exportado correctamente.")
                                    else:
                                        print("No hay reservaciones registradas para exportar.")

                            except Error as e:
                                print(f"Error de Base de Datos: {e}")
                            except Exception as ex:
                                print(f"Error inesperado al exportar a Excel: {ex}")

                        else:
                            print('Formato no valido')
                    elif opcion_exportar == 'N':
                        print('Regresando...')
                        break
                    else:
                        print('Ingrese una opcion valida')
                        continue
            else:
                print(f'No hay reservaciones hechas para el dia {fecha_consulta}')




    if opcion == '4':
        
        while True:
            try:
                fecha_inicio_str = input("\nIngrese fecha de inicio (MM-DD-AAAA) ('X' para salir): ").strip()
                if fecha_inicio_str.upper() == 'X':
                    print('Saliendo...')
                    break
                fecha_fin_str = input("Ingrese fecha final (MM-DD-AAAA): ").strip()
                
                fecha_inicio = datetime.strptime(fecha_inicio_str, "%m-%d-%Y").date()
                fecha_fin = datetime.strptime(fecha_fin_str, "%m-%d-%Y").date()
                if fecha_fin < fecha_inicio:
                    print("Error. La fecha final no puede ser anterior a la fecha de inicio.")
                    continue
            except ValueError:
                print("Formato de fecha inválido")
                continue

            fechas = (fecha_inicio, fecha_fin)

            try:
                with sqlite3.connect('espacios_coworking.db') as conn:
                    cursor = conn.cursor()
                    cursor.execute('SELECT folio_reservacion, nombre_reservacion, \
                                    turno_reservacion, fecha_reservacion \
                                    FROM reservaciones \
                                    WHERE fecha_reservacion BETWEEN ? AND ? \
                                    AND estatus_reservacion = "Activa"', fechas)
                    reservaciones = cursor.fetchall()
            except Error as e:
                print(f"Error de base de datos: {e}")
                continue

            if not reservaciones:
                print("No hay reservaciones en este rango de fechas.")
                continue


            print('\n***** RESERVACIONES ACTIVAS EN EL RANGO DE FECHAS *****')
            print(f'{"FOLIO":<10}{"NOMBRE":<25}{"TURNO":<15}{"FECHA"}')
            print('='*55)
            for folio, nombre, turno, fecha in reservaciones:
                print(f'{folio:<10}{nombre:<25}{turno:<15}{fecha}')
            
            while True:
                folio_cancelar = input("\nIngrese el folio que desea cancelar ('X' para salir): ").strip().upper()
                if folio_cancelar == 'X':
                    break
                if not folio_cancelar.isdigit():
                    print("Folio inválido.")
                    continue
                if int(folio_cancelar) not in [r[0] for r in reservaciones]:
                    print("Folio no encontrado en las reservaciones activas mostradas.")
                    continue
                break

            if folio_cancelar == 'X':
                continue

            folio_cancelar = int(folio_cancelar)
            
            try:
                with sqlite3.connect('espacios_coworking.db') as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT fecha_reservacion FROM reservaciones WHERE folio_reservacion = ?", (folio_cancelar,))
                    fecha_res = cursor.fetchone()
            except Error as e:
                print(f"Error de base de datos: {e}")
                continue

            if not fecha_res:
                print("Folio no encontrado.")
                continue
            
            fecha_reservada = datetime.fromisoformat(fecha_res[0]).date() 

            if (fecha_reservada - hoy).days < 2:
                print("Solo se puede cancelar con al menos 2 días de anticipación.")
                continue

            confirmar = input(f"Confirma la cancelacion de la reservación {folio_cancelar}? (S/N): ").upper().strip()
            if confirmar == 'S':
                try:
                    with sqlite3.connect('espacios_coworking.db') as conn:
                        cursor = conn.cursor()
                        cursor.execute("UPDATE reservaciones SET estatus_reservacion = 'Cancelada' \
                                       WHERE folio_reservacion = ?", (folio_cancelar,))
                        print("Reservación cancelada correctamente.")
                except Error as e:
                    print(f"Error de base de datos: {e}")
                    continue
            else:
                print("Operación cancelada por el usuario.")
            
            break
            



    if opcion == '5':
        nombre_cliente = input('\nIngrese su nombre ("X" para salir): ').strip().title()
        if nombre_cliente == 'X':
            print('Saliendo...')
            continue
        apellido_cliente = input('Ingrese su apellido: ').strip().title()

        if nombre_cliente and apellido_cliente:

            clientes = (nombre_cliente, apellido_cliente)
                
            try:
                with sqlite3.connect('espacios_coworking.db') as conn:
                    cursor = conn.cursor()
                    cursor.execute('INSERT INTO clientes (nombre, apellido) values(?,?)', clientes)
                    print('Cliente registrado')
            except Error as e:
                print(e)
            except:
                print(f'Se produjo un error: {sys.exc_info()[0]}')
        else:
            print('Error. El nombre y apellido no pueden estar vacios') 



    if opcion == '6':
        nombre_sala = input('\nIngrese el nombre de su sala ("X" para salir): ').strip().title()
        if nombre_sala == 'X':
            print('Saliendo...')
            continue
        cupo_sala = input('Ingrese el cupo de su sala: ').strip() 

        if nombre_sala and cupo_sala:
            try:
                cupo_sala  = int(cupo_sala)
                if cupo_sala <= 0:
                    raise ValueError
            except ValueError:
                print('Error. Cupo de sala invalido')
            else:
                salas = (nombre_sala, cupo_sala)
                try:
                    with sqlite3.connect('espacios_coworking.db') as conn:
                        cursor = conn.cursor()
                        cursor.execute('INSERT INTO salas (nombre_sala, cupo_sala) values(?,?)', salas)
                        print('Sala registrada')
                except Error as e:
                    print(e)
                except:
                    print(f'Se produjo un error: {sys.exc_info()[0]}')
        else:
            print('Error. El nombre y cupo no pueden estar vacios')



    if opcion == '7':
        confirmar_salida = input('Desea confirmar su salida? (S/N): ').upper().strip()
        if confirmar_salida == 'S':
            print('Gracias por su visita!')
            break
        elif confirmar_salida == 'N':
            print('Regresando...')
            continue
        else:
            print('Opción no valida')
            continue